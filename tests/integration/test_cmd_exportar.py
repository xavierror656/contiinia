"""Pruebas de integración para `contiinia exportar`.

Cubre CA-EXP-01..09.
Usa typer CliRunner y subprocess según el caso.
El directorio de fixtures existente contiene XMLs de CFDI.
"""

from __future__ import annotations

import json
import os
import stat
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest
from typer.testing import CliRunner

from contiinia.cli import app

runner = CliRunner()

FIXTURES = Path(__file__).parent.parent.parent / "fixtures"

# Columnas esperadas en Detalle_CFDI según spec 10.4
COLUMNAS_DETALLE = [
    "uuid", "tipo", "fecha", "emisor_rfc", "emisor_nombre",
    "receptor_rfc", "receptor_nombre", "moneda",
    "subtotal", "total", "metodo_pago", "forma_pago",
]

HOJAS_ESPERADAS = ["Resumen", "Detalle_CFDI", "Errores"]


def _run_exportar(args: list[str]) -> tuple[int, str, str]:
    """Ejecuta contiinia exportar por subprocess y devuelve (exit_code, stdout, stderr)."""
    resultado = subprocess.run(
        [sys.executable, "-m", "contiinia.cli"] + args,
        capture_output=True,
        text=True,
    )
    return resultado.returncode, resultado.stdout, resultado.stderr


# ---------------------------------------------------------------------------
# CA-EXP-01: directorio con XMLs válidos → exit 0, .xlsx generado, stdout JSON
# ---------------------------------------------------------------------------


def test_exportar_exit_0_con_xmls(tmp_path: Path) -> None:
    """CA-EXP-01: directorio con XMLs válidos → exit 0."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert result.exit_code == 0, f"Esperaba exit 0, output: {result.output}"


def test_exportar_xlsx_generado(tmp_path: Path) -> None:
    """CA-EXP-01: el archivo .xlsx es creado en disco."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert salida.exists(), "El archivo .xlsx no fue creado"


def test_exportar_stdout_es_json_valido(tmp_path: Path) -> None:
    """CA-EXP-01: stdout es JSON parseable."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    assert isinstance(data, dict)


def test_exportar_stdout_campos_requeridos(tmp_path: Path) -> None:
    """CA-EXP-01: stdout contiene archivo_generado, hojas, total_cfdi_exitosos, total_cfdi_errores, periodo."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    for campo in ("archivo_generado", "hojas", "total_cfdi_exitosos", "total_cfdi_errores", "periodo"):
        assert campo in data, f"Falta campo: {campo}"


def test_exportar_total_exitosos_entero(tmp_path: Path) -> None:
    """CA-EXP-01: total_cfdi_exitosos es entero, no string."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    assert isinstance(data["total_cfdi_exitosos"], int), (
        f"total_cfdi_exitosos debe ser int, es: {type(data['total_cfdi_exitosos'])}"
    )


def test_exportar_total_errores_entero(tmp_path: Path) -> None:
    """CA-EXP-01: total_cfdi_errores es entero, no string."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    assert isinstance(data["total_cfdi_errores"], int)


# ---------------------------------------------------------------------------
# CA-EXP-02: hojas correctas, 3 en total
# ---------------------------------------------------------------------------


def test_exportar_xlsx_tiene_3_hojas(tmp_path: Path) -> None:
    """CA-EXP-02: el .xlsx tiene exactamente 3 hojas."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    assert len(wb.sheetnames) == 3, f"Esperaba 3 hojas, encontré: {wb.sheetnames}"


def test_exportar_xlsx_nombres_hojas_correctos(tmp_path: Path) -> None:
    """CA-EXP-02: las 3 hojas se llaman Resumen, Detalle_CFDI, Errores en ese orden."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    assert wb.sheetnames == HOJAS_ESPERADAS, f"Hojas: {wb.sheetnames}"


def test_exportar_stdout_hojas_array(tmp_path: Path) -> None:
    """CA-EXP-02: stdout['hojas'] es ['Resumen', 'Detalle_CFDI', 'Errores']."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    assert data["hojas"] == HOJAS_ESPERADAS


def test_exportar_detalle_cfdi_fila_por_exitoso(tmp_path: Path) -> None:
    """CA-EXP-02: Detalle_CFDI tiene una fila por CFDI exitoso (excluyendo encabezado)."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    n_exitosos = data["total_cfdi_exitosos"]

    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Detalle_CFDI"]
    filas_datos = ws.max_row - 1  # excluir encabezado
    assert filas_datos == n_exitosos, (
        f"Detalle_CFDI tiene {filas_datos} filas de datos, pero se reportaron {n_exitosos} exitosos"
    )


def test_exportar_errores_fila_por_error(tmp_path: Path) -> None:
    """CA-EXP-02: hoja Errores tiene una fila por CFDI con error (excluyendo encabezado)."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    n_errores = data["total_cfdi_errores"]

    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Errores"]
    filas_datos = ws.max_row - 1
    assert filas_datos == n_errores, (
        f"Errores tiene {filas_datos} filas, se reportaron {n_errores} errores"
    )


# ---------------------------------------------------------------------------
# CA-EXP-03: Hoja Resumen contiene columnas correctas con valores
# ---------------------------------------------------------------------------


def test_exportar_resumen_hoja_tiene_filas(tmp_path: Path) -> None:
    """CA-EXP-03: hoja Resumen tiene al menos las filas de subtotal_ingresos y total_neto."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Resumen"]
    claves = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "subtotal_ingresos" in claves, f"Falta subtotal_ingresos en Resumen. Claves: {claves}"
    assert "total_neto" in claves, f"Falta total_neto en Resumen. Claves: {claves}"


def test_exportar_resumen_todos_campos_esperados(tmp_path: Path) -> None:
    """CA-EXP-03: hoja Resumen tiene todos los campos de la spec 10.4."""
    campos_esperados = [
        "subtotal_ingresos", "subtotal_egresos", "total_neto",
        "total_iva_trasladado", "total_iva_retenido", "total_isr_retenido",
        "fecha_min", "fecha_max",
        "count_I", "count_E", "count_P", "count_N", "count_T",
    ]
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Resumen"]
    claves = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    for campo in campos_esperados:
        assert campo in claves, f"Falta campo '{campo}' en hoja Resumen. Claves encontradas: {claves}"


def test_exportar_resumen_importes_son_strings(tmp_path: Path) -> None:
    """CA-EXP-03: los importes en Resumen son strings, no celdas numéricas."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Resumen"]
    campos_monetarios = {
        "subtotal_ingresos", "subtotal_egresos", "total_neto",
        "total_iva_trasladado", "total_iva_retenido", "total_isr_retenido",
    }
    for r in range(1, ws.max_row + 1):
        clave = ws.cell(row=r, column=1).value
        valor = ws.cell(row=r, column=2).value
        if clave in campos_monetarios:
            assert isinstance(valor, str), (
                f"Campo '{clave}' en Resumen debe ser string, es {type(valor)}: {valor!r}"
            )


# ---------------------------------------------------------------------------
# CA-EXP-03 cont.: Detalle_CFDI tiene las 12 columnas correctas
# ---------------------------------------------------------------------------


def test_exportar_detalle_tiene_12_columnas(tmp_path: Path) -> None:
    """CA-EXP-03: Detalle_CFDI tiene exactamente 12 columnas de encabezado."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Detalle_CFDI"]
    encabezados = [ws.cell(row=1, column=c).value for c in range(1, 13)]
    assert encabezados == COLUMNAS_DETALLE, f"Encabezados incorrectos: {encabezados}"


def test_exportar_errores_tiene_3_columnas(tmp_path: Path) -> None:
    """CA-EXP-02: hoja Errores tiene columnas archivo, error, detalle."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Errores"]
    encabezados = [ws.cell(row=1, column=c).value for c in range(1, 4)]
    assert encabezados == ["archivo", "error", "detalle"], f"Encabezados de Errores: {encabezados}"


# ---------------------------------------------------------------------------
# CA-EXP-04: extensión distinta de .xlsx → exit 1, error formato_no_soportado
# ---------------------------------------------------------------------------


def test_exportar_salida_xls_exit_1(tmp_path: Path) -> None:
    """CA-EXP-04: salida.xls → exit 1."""
    salida = tmp_path / "reporte.xls"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert result.exit_code == 1, f"Esperaba exit 1, output: {result.output}"


def test_exportar_salida_csv_exit_1(tmp_path: Path) -> None:
    """CA-EXP-04: salida.csv → exit 1."""
    salida = tmp_path / "reporte.csv"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert result.exit_code == 1


def test_exportar_salida_xls_no_genera_archivo(tmp_path: Path) -> None:
    """CA-EXP-04: con extensión incorrecta, NO se genera ningún archivo."""
    salida = tmp_path / "reporte.xls"
    runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert not salida.exists(), "No debería generarse el archivo con extensión incorrecta"


def test_exportar_salida_xls_stderr_error(tmp_path: Path) -> None:
    """CA-EXP-04: stderr contiene JSON con error='formato_no_soportado'."""
    salida = tmp_path / "reporte.xls"
    exit_code, stdout, stderr = _run_exportar(["exportar", str(FIXTURES), str(salida)])
    assert exit_code == 1
    err = json.loads(stderr)
    assert err["error"] == "formato_no_soportado"


# ---------------------------------------------------------------------------
# CA-EXP-05: directorio no existe → exit 3, error directorio_no_encontrado
# ---------------------------------------------------------------------------


def test_exportar_directorio_no_existe_exit_3(tmp_path: Path) -> None:
    """CA-EXP-05: directorio no existe → exit 3."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", "/tmp/directorio_que_no_existe_xyz123", str(salida)])
    assert result.exit_code == 3, f"output: {result.output}"


def test_exportar_directorio_no_existe_stderr(tmp_path: Path) -> None:
    """CA-EXP-05: stderr JSON con error='directorio_no_encontrado'."""
    salida = tmp_path / "reporte.xlsx"
    exit_code, stdout, stderr = _run_exportar(
        ["exportar", "/tmp/directorio_que_no_existe_xyz123", str(salida)]
    )
    assert exit_code == 3
    err = json.loads(stderr)
    assert err["error"] == "directorio_no_encontrado"


def test_exportar_directorio_no_existe_no_genera_archivo(tmp_path: Path) -> None:
    """CA-EXP-05: con directorio inexistente, no se genera ningún archivo."""
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", "/tmp/directorio_que_no_existe_xyz123", str(salida)])
    assert not salida.exists()


# ---------------------------------------------------------------------------
# CA-EXP-06: ruta de salida en directorio sin permisos → exit 3, error_escritura
# ---------------------------------------------------------------------------


@pytest.mark.skipif(os.getuid() == 0, reason="Root ignora permisos de escritura")
def test_exportar_sin_permisos_escritura_exit_3(tmp_path: Path) -> None:
    """CA-EXP-06: directorio de salida sin permisos → exit 3."""
    dir_sin_permisos = tmp_path / "readonly"
    dir_sin_permisos.mkdir()
    dir_sin_permisos.chmod(stat.S_IRUSR | stat.S_IXUSR)  # solo lectura y ejecución
    try:
        salida = dir_sin_permisos / "reporte.xlsx"
        result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
        assert result.exit_code == 3, f"Esperaba exit 3, output: {result.output}"
    finally:
        dir_sin_permisos.chmod(stat.S_IRWXU)  # restaurar para limpieza


@pytest.mark.skipif(os.getuid() == 0, reason="Root ignora permisos de escritura")
def test_exportar_sin_permisos_no_genera_archivo(tmp_path: Path) -> None:
    """CA-EXP-06: sin permisos de escritura, no queda archivo parcial."""
    dir_sin_permisos = tmp_path / "readonly"
    dir_sin_permisos.mkdir()
    dir_sin_permisos.chmod(stat.S_IRUSR | stat.S_IXUSR)
    try:
        salida = dir_sin_permisos / "reporte.xlsx"
        runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
        assert not salida.exists()
    finally:
        dir_sin_permisos.chmod(stat.S_IRWXU)


# ---------------------------------------------------------------------------
# CA-EXP-07: directorio vacío (sin XML) → exit 0, xlsx con solo encabezados
# ---------------------------------------------------------------------------


def test_exportar_directorio_vacio_exit_0(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → exit 0."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    assert result.exit_code == 0, f"output: {result.output}"


def test_exportar_directorio_vacio_xlsx_generado(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → el .xlsx es creado."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    assert salida.exists()


def test_exportar_directorio_vacio_exitosos_0(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → total_cfdi_exitosos=0."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    data = json.loads(result.output)
    assert data["total_cfdi_exitosos"] == 0


def test_exportar_directorio_vacio_errores_0(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → total_cfdi_errores=0."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    data = json.loads(result.output)
    assert data["total_cfdi_errores"] == 0


def test_exportar_directorio_vacio_fecha_min_null(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → periodo.fecha_min=null."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    data = json.loads(result.output)
    assert data["periodo"]["fecha_min"] is None, f"fecha_min debe ser null: {data['periodo']}"


def test_exportar_directorio_vacio_fecha_max_null(tmp_path: Path) -> None:
    """CA-EXP-07: directorio vacío → periodo.fecha_max=null."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    data = json.loads(result.output)
    assert data["periodo"]["fecha_max"] is None


def test_exportar_directorio_vacio_detalle_solo_encabezado(tmp_path: Path) -> None:
    """CA-EXP-07: Detalle_CFDI con directorio vacío → solo encabezado, sin filas de datos."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Detalle_CFDI"]
    assert ws.max_row == 1, f"Detalle_CFDI debe tener solo encabezado, tiene {ws.max_row} filas"


def test_exportar_directorio_vacio_errores_solo_encabezado(tmp_path: Path) -> None:
    """CA-EXP-07: hoja Errores con directorio vacío → solo encabezado."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Errores"]
    assert ws.max_row == 1, f"Errores debe tener solo encabezado, tiene {ws.max_row} filas"


def test_exportar_directorio_vacio_resumen_totales_cero(tmp_path: Path) -> None:
    """CA-EXP-07: Resumen con directorio vacío → subtotal_ingresos='0', total_neto='0'."""
    dir_vacio = tmp_path / "vacio"
    dir_vacio.mkdir()
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(dir_vacio), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws = wb["Resumen"]
    vals = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)}
    from decimal import Decimal
    assert Decimal(str(vals["subtotal_ingresos"])) == Decimal("0")
    assert Decimal(str(vals["total_neto"])) == Decimal("0")


# ---------------------------------------------------------------------------
# CA-EXP-08: archivo_generado es ruta absoluta
# ---------------------------------------------------------------------------


def test_exportar_archivo_generado_es_absoluto(tmp_path: Path) -> None:
    """CA-EXP-08: archivo_generado en stdout es ruta absoluta."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    ruta_gen = Path(data["archivo_generado"])
    assert ruta_gen.is_absolute(), f"archivo_generado no es absoluto: {data['archivo_generado']}"


def test_exportar_archivo_generado_coincide_con_archivo(tmp_path: Path) -> None:
    """CA-EXP-08: archivo_generado apunta al mismo archivo xlsx creado."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    data = json.loads(result.output)
    ruta_gen = Path(data["archivo_generado"])
    assert ruta_gen.exists(), f"El archivo en archivo_generado no existe: {ruta_gen}"


# ---------------------------------------------------------------------------
# CA-EXP-09: --schema → exit 0, JSON Schema, sin leer directorio
# ---------------------------------------------------------------------------


def test_exportar_schema_exit_0(tmp_path: Path) -> None:
    """CA-EXP-09: --schema → exit 0."""
    # Usamos un directorio que no existe; --schema no debe tocarlo
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", "--schema", "/no/existe/directorio", str(salida)])
    assert result.exit_code == 0, f"output: {result.output}"


def test_exportar_schema_es_json_valido(tmp_path: Path) -> None:
    """CA-EXP-09: --schema emite JSON parseable."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", "--schema", str(FIXTURES), str(salida)])
    schema = json.loads(result.output)
    assert isinstance(schema, dict)


def test_exportar_schema_tiene_estructura_json_schema(tmp_path: Path) -> None:
    """CA-EXP-09: --schema emite JSON Schema válido (tiene 'properties' o '$defs' o 'title')."""
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", "--schema", str(FIXTURES), str(salida)])
    schema = json.loads(result.output)
    assert "properties" in schema or "$defs" in schema or "title" in schema


def test_exportar_schema_no_genera_archivo(tmp_path: Path) -> None:
    """CA-EXP-09: --schema no genera ningún archivo de salida."""
    salida = tmp_path / "reporte_schema.xlsx"
    runner.invoke(app, ["exportar", "--schema", str(FIXTURES), str(salida)])
    assert not salida.exists(), "--schema no debe generar archivo xlsx"


# ---------------------------------------------------------------------------
# Adversarial: la ruta del padre de salida no existe → exit 3, error_escritura
# ---------------------------------------------------------------------------


def test_exportar_directorio_padre_no_existe_exit_3(tmp_path: Path) -> None:
    """Adversarial CA-EXP-06: si el directorio padre del .xlsx no existe → exit 3."""
    salida = tmp_path / "subdir_inexistente" / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(FIXTURES), str(salida)])
    assert result.exit_code == 3, f"Esperaba exit 3, output: {result.output}"


# ---------------------------------------------------------------------------
# Adversarial: CFDI 3.3 en el directorio cuenta como error
# ---------------------------------------------------------------------------


def test_exportar_cfdi_33_cuenta_como_error(tmp_path: Path) -> None:
    """Adversarial: un directorio con solo el CFDI 3.3 → total_cfdi_errores>=1, total_cfdi_exitosos=0."""
    dir_solo_33 = tmp_path / "solo33"
    dir_solo_33.mkdir()
    import shutil
    shutil.copy(str(FIXTURES / "cfdi_3.3_rechazado.xml"), str(dir_solo_33 / "cfdi_33.xml"))
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_solo_33), str(salida)])
    assert result.exit_code == 0  # El proceso no aborta
    data = json.loads(result.output)
    assert data["total_cfdi_exitosos"] == 0
    assert data["total_cfdi_errores"] >= 1


def test_exportar_cfdi_33_en_hoja_errores(tmp_path: Path) -> None:
    """Adversarial: CFDI 3.3 aparece en hoja Errores, no en Detalle_CFDI."""
    dir_solo_33 = tmp_path / "solo33"
    dir_solo_33.mkdir()
    import shutil
    shutil.copy(str(FIXTURES / "cfdi_3.3_rechazado.xml"), str(dir_solo_33 / "cfdi_33.xml"))
    salida = tmp_path / "reporte.xlsx"
    runner.invoke(app, ["exportar", str(dir_solo_33), str(salida)])
    wb = openpyxl.load_workbook(str(salida))
    ws_err = wb["Errores"]
    ws_det = wb["Detalle_CFDI"]
    assert ws_err.max_row > 1, "CFDI 3.3 debe aparecer en hoja Errores"
    assert ws_det.max_row == 1, "Detalle_CFDI solo debe tener encabezado cuando hay 0 exitosos"


# ---------------------------------------------------------------------------
# Adversarial: verificar que archivos no-xml en el directorio no se procesan
# ---------------------------------------------------------------------------


def test_exportar_ignora_archivos_no_xml(tmp_path: Path) -> None:
    """Adversarial: directorio con archivos .csv y .txt → no se intentan parsear como CFDI."""
    dir_mixto = tmp_path / "mixto"
    dir_mixto.mkdir()
    (dir_mixto / "archivo.csv").write_text("dato,dato\n")
    (dir_mixto / "archivo.txt").write_text("texto\n")
    salida = tmp_path / "reporte.xlsx"
    result = runner.invoke(app, ["exportar", str(dir_mixto), str(salida)])
    assert result.exit_code == 0
    data = json.loads(result.output)
    assert data["total_cfdi_exitosos"] == 0
    assert data["total_cfdi_errores"] == 0
