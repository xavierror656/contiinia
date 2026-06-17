"""Pruebas unitarias para parsers/tabla.py — cubre CA-TAB-01..09."""

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import pytest

from contiinia.errors import ArchivoNoEncontradoError, FormatoNoSoportadoError
from contiinia.parsers.tabla import parsear_tabla

FIXTURES = Path(__file__).parent.parent.parent / "fixtures"


def fx(nombre: str) -> Path:
    return FIXTURES / nombre


# ---------------------------------------------------------------------------
# CA-TAB-01: CSV básico con columnas canónicas
# ---------------------------------------------------------------------------


def test_csv_basico_filas() -> None:
    """CA-TAB-01: CSV con columnas canónicas → 2 filas."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    assert result.total_registros == 2


def test_csv_basico_columnas_detectadas() -> None:
    """CA-TAB-01: columnas_detectadas incluye 'clave_prod_serv'."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    assert "clave_prod_serv" in result.columnas_detectadas


def test_csv_basico_registro_clave() -> None:
    """CA-TAB-01: primer registro tiene clave_prod_serv='81161500'."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    assert result.registros[0].clave_prod_serv == "81161500"


def test_csv_basico_registro_descripcion() -> None:
    """CA-TAB-01: primer registro tiene descripcion correcta."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    assert result.registros[0].descripcion == "Consultoria"


# ---------------------------------------------------------------------------
# CA-TAB-03: XLSX básico — mismo esquema que CSV
# ---------------------------------------------------------------------------


def test_xlsx_basico_filas() -> None:
    """CA-TAB-03: XLSX válido → mismo número de filas que CSV."""
    result = parsear_tabla(fx("tabla_conceptos.xlsx"))
    assert result.total_registros == 2


def test_xlsx_basico_columnas_detectadas() -> None:
    """CA-TAB-03: XLSX columnas_detectadas incluye 'clave_prod_serv'."""
    result = parsear_tabla(fx("tabla_conceptos.xlsx"))
    assert "clave_prod_serv" in result.columnas_detectadas


def test_xlsx_basico_registro_clave() -> None:
    """CA-TAB-03: primer registro XLSX tiene clave_prod_serv correcta."""
    result = parsear_tabla(fx("tabla_conceptos.xlsx"))
    assert result.registros[0].clave_prod_serv == "81161500"


# ---------------------------------------------------------------------------
# CA-TAB-08: Decimales como string en JSON
# ---------------------------------------------------------------------------


def test_decimal_como_string_en_json() -> None:
    """CA-TAB-08: cantidad, valor_unitario, importe son strings en JSON."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    json_str = result.model_dump_json()
    data = json.loads(json_str)
    reg = data["registros"][0]
    assert isinstance(reg["cantidad"], str)
    assert isinstance(reg["valor_unitario"], str)
    assert isinstance(reg["importe"], str)


def test_no_floats_en_json() -> None:
    """CA-TAB-08: ningún valor en JSON es float (type number)."""
    result = parsear_tabla(fx("tabla_conceptos.csv"))
    data = json.loads(result.model_dump_json())

    def check_no_float(obj: object, path: str = "") -> None:
        if isinstance(obj, dict):
            for k, v in obj.items():
                check_no_float(v, f"{path}.{k}")
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                check_no_float(v, f"{path}[{i}]")
        elif isinstance(obj, float):
            raise AssertionError(f"Float encontrado en {path}: {obj!r}")

    check_no_float(data)


# ---------------------------------------------------------------------------
# CA-TAB-04: Alias de columnas
# ---------------------------------------------------------------------------


def test_alias_precio_mapeado_a_valor_unitario() -> None:
    """CA-TAB-04: columna 'precio' es mapeada a valor_unitario."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,descripcion,cantidad,precio,importe\n")
        f.write("84111506,Servicio,1,500.00,500.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert "valor_unitario" in result.columnas_detectadas
    assert result.registros[0].valor_unitario == Decimal("500.00")


def test_alias_concepto_mapeado_a_descripcion() -> None:
    """CA-TAB-04: columna 'concepto' mapeada a descripcion."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,concepto,cantidad,valor_unitario,importe\n")
        f.write("84111506,Consultoria,2,100.00,200.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert "descripcion" in result.columnas_detectadas
    assert result.registros[0].descripcion == "Consultoria"


def test_alias_qty_mapeado_a_cantidad() -> None:
    """CA-TAB-04: columna 'qty' mapeada a cantidad."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,descripcion,qty,valor_unitario,importe\n")
        f.write("84111506,Servicio,3,100.00,300.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert "cantidad" in result.columnas_detectadas
    assert result.registros[0].cantidad == Decimal("3")


def test_alias_monto_mapeado_a_importe() -> None:
    """CA-TAB-04: columna 'monto' mapeada a importe."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,descripcion,cantidad,valor_unitario,monto\n")
        f.write("84111506,Servicio,1,250.00,250.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert "importe" in result.columnas_detectadas
    assert result.registros[0].importe == Decimal("250.00")


# ---------------------------------------------------------------------------
# CA-TAB-07: Columnas extra van en extras
# ---------------------------------------------------------------------------


def test_columnas_extra_en_extras() -> None:
    """CA-TAB-07: columnas no canónicas van en extras sin causar error."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,descripcion,cantidad,valor_unitario,importe,referencia\n")
        f.write("84111506,Servicio,1,100.00,100.00,REF-001\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert result.registros[0].columnas_extra.get("referencia") == "REF-001"


# ---------------------------------------------------------------------------
# CA-TAB-02: CSV con delimitador punto y coma
# ---------------------------------------------------------------------------


def test_csv_semicolon_detectado() -> None:
    """CA-TAB-02: CSV con punto y coma como separador → detectado automáticamente."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv;descripcion;cantidad;valor_unitario;importe\n")
        f.write("84111506;Servicio;1;100.00;100.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert result.total_registros == 1
    assert result.registros[0].clave_prod_serv == "84111506"


# ---------------------------------------------------------------------------
# Archivo no existe → ArchivoNoEncontradoError (exit 3)
# ---------------------------------------------------------------------------


def test_archivo_no_existente_raise_error() -> None:
    """CA-TAB (error): archivo no existente → ArchivoNoEncontradoError."""
    with pytest.raises(ArchivoNoEncontradoError):
        parsear_tabla(Path("/tmp/no_existe_tabla_xyz.csv"))


# ---------------------------------------------------------------------------
# Extensión no soportada → FormatoNoSoportadoError (exit 1)
# ---------------------------------------------------------------------------


def test_extension_no_soportada_raise_error() -> None:
    """CA-TAB (error): extensión .txt → FormatoNoSoportadoError."""
    with tempfile.NamedTemporaryFile(suffix=".txt", mode="w", delete=False, encoding="utf-8") as f:
        f.write("dato\n")
        tmp = Path(f.name)

    with pytest.raises(FormatoNoSoportadoError):
        parsear_tabla(tmp)

    tmp.unlink()


# ---------------------------------------------------------------------------
# CA-TAB-06: Filas vacías omitidas
# ---------------------------------------------------------------------------


def test_filas_vacias_omitidas() -> None:
    """CA-TAB-06: filas completamente vacías no se incluyen en registros."""
    with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", delete=False, encoding="utf-8") as f:
        f.write("clave_prod_serv,descripcion,cantidad,valor_unitario,importe\n")
        f.write("84111506,Servicio,1,100.00,100.00\n")
        f.write(",,,\n")  # fila vacía
        f.write("84111507,Otro,2,50.00,100.00\n")
        tmp = Path(f.name)

    result = parsear_tabla(tmp)
    tmp.unlink()

    assert result.total_registros == 2


# ---------------------------------------------------------------------------
# QA-TAB-01: XLSX multi-hoja → primera hoja procesada + advertencia
# ---------------------------------------------------------------------------


def test_xlsx_multi_hoja_procesa_primera(tmp_path: Path) -> None:
    """QA-TAB-01: XLSX con múltiples hojas → procesa primera, emite advertencia."""
    import openpyxl
    ruta = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Hoja1"
    ws1.append(["clave_prod_serv", "descripcion", "cantidad", "valor_unitario", "importe"])
    ws1.append(["84111506", "Servicio", "1", "100.00", "100.00"])
    ws2 = wb.create_sheet("Hoja2")
    ws2.append(["clave_prod_serv", "descripcion", "cantidad", "valor_unitario", "importe"])
    ws2.append(["99999999", "Otro", "5", "200.00", "1000.00"])
    wb.save(ruta)

    result = parsear_tabla(ruta)

    # Solo la primera hoja: 1 registro
    assert result.total_registros == 1
    assert result.registros[0].clave_prod_serv == "84111506"
    # Advertencia presente
    assert len(result.advertencias) >= 1
    assert "Hoja1" in result.advertencias[0]


def test_xlsx_una_hoja_sin_advertencia(tmp_path: Path) -> None:
    """QA-TAB-01: XLSX con una sola hoja → sin advertencia."""
    import openpyxl
    ruta = tmp_path / "single.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conceptos"
    ws.append(["clave_prod_serv", "descripcion", "cantidad", "valor_unitario", "importe"])
    ws.append(["84111506", "Servicio", "1", "100.00", "100.00"])
    wb.save(ruta)

    result = parsear_tabla(ruta)

    assert result.total_registros == 1
    assert result.advertencias == []
