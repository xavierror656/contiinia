"""Script auxiliar para generar los fixtures CONTPAQi. Ejecutar una sola vez."""
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent


def gen_comercial_movimientos() -> None:
    """Simula el export 'Listado de Movimientos' de CONTPAQi Comercial Start/Pro."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Encabezados tal como los genera CONTPAQi Comercial
    ws.append(["Código", "Nombre", "Unidades", "Precio Unit", "Monto", "% IVA"])

    rows = [
        # (Código, Nombre, Unidades, Precio Unit, Monto, % IVA)
        ["84111506", "Consultoría en tecnología", 1, "2500.00", "2500.00", "0.16"],
        ["84111507", "Soporte técnico mensual",   1, "1800.00", "1800.00", "0.16"],
        ["81111501", "Capacitación en línea",      2,  "750.00", "1500.00", "0.16"],
        ["01010101", "Servicios médicos",           1, "3200.00", "3200.00", "0"],
        ["50171544", "Papel bond carta 500 hojas", 10,   "85.00",  "850.00", "0.16"],
        ["44121618", "Tóner impresora láser",       2, "1250.00", "2500.00", ""],
    ]
    for r in rows:
        ws.append(r)

    out = BASE / "comercial_movimientos.xlsx"
    wb.save(out)
    print(f"Creado: {out}")


def gen_factura_electronica_conceptos() -> None:
    """Simula el reporte de conceptos de CONTPAQi Factura Electrónica."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conceptos"

    # Encabezados camelCase tal como los exporta CONTPAQi Factura Electrónica
    ws.append(["ClaveProdServ", "Descripcion", "Cantidad", "ValorUnitario", "Importe", "TasaOCuota"])

    rows = [
        ["84111506", "Desarrollo de software a la medida",   1, "12000.00", "12000.00", "0.16"],
        ["84111507", "Mantenimiento de sistemas",             1,  "4500.00",  "4500.00", "0.16"],
        ["81111501", "Curso de capacitación ERP",             3,  "1200.00",  "3600.00", "0.16"],
        ["01010101", "Honorarios médicos especialista",       1,  "2800.00",  "2800.00", "0"],
        ["50202300", "Licencia de software anual",            2,  "6000.00", "12000.00", "0.16"],
    ]
    for r in rows:
        ws.append(r)

    out = BASE / "factura_electronica_conceptos.xlsx"
    wb.save(out)
    print(f"Creado: {out}")


def gen_add_hoja_electronica() -> None:
    """Simula la hoja electrónica del ADD (Administrador de Documentos Digitales)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja Electrónica"

    # Encabezados ADD — IVA en lugar de TasaOCuota
    ws.append(["ClaveProdServ", "Descripcion", "Cantidad", "ValorUnitario", "Importe", "IVA"])

    rows = [
        # IVA: valores mixtos (0.16, 0.08, "Exento") para probar mapeo a tasa
        ["84111506", "Asesoría fiscal y contable",    1,  "5000.00",  "5000.00", "0.16"],
        ["84111507", "Auditoría de estados financ.",  1,  "8500.00",  "8500.00", "0.16"],
        ["01010101", "Servicios educativos privados",  5,   "650.00",  "3250.00", "Exento"],
        ["80141600", "Arrendamiento de oficina",       1,  "3200.00",  "3200.00", "0.16"],
        ["50202300", "Software región fronteriza",     1,  "4000.00",  "4000.00", "0.08"],
    ]
    for r in rows:
        ws.append(r)

    out = BASE / "add_hoja_electronica.xlsx"
    wb.save(out)
    print(f"Creado: {out}")


if __name__ == "__main__":
    gen_comercial_movimientos()
    gen_factura_electronica_conceptos()
    gen_add_hoja_electronica()
    print("Listo.")
