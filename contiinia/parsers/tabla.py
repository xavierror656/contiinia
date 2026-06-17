"""Parser de tablas de conceptos CSV/XLSX — CA-TAB-01..09."""

from __future__ import annotations

import csv
import io
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import xlrd
from lxml.etree import fromstring as etree_fromstring
from openpyxl import load_workbook

from contiinia.errors import (
    ArchivoCorrumpidoError,
    ArchivoNoEncontradoError,
    ArchivoSinDatosError,
    ColumnaRequeridaAusenteError,
    FormatoNoSoportadoError,
    ValorNoNumericoError,
)
from contiinia.models.tabla import (
    AdvertenciaImporteInconsistente,
    AdvertenciaTabla,
    AdvertenciaTasaNoNumerica,
    TablaResult,
    TablaRow,
)

# ---------------------------------------------------------------------------
# Constantes de namespace ODS
# ---------------------------------------------------------------------------

_NS_TABLE = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
_NS_TEXT = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
_NS_OFFICE = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"

# ---------------------------------------------------------------------------
# Columnas requeridas según spec 4.5
# ---------------------------------------------------------------------------

_COLUMNAS_REQUERIDAS = frozenset(
    {"clave_prod_serv", "descripcion", "cantidad", "valor_unitario", "importe"}
)

# ---------------------------------------------------------------------------
# Mapeo alias → columna canónica
# ---------------------------------------------------------------------------

_ALIAS_MAP: dict[str, str] = {}

_ALIASES: dict[str, list[str]] = {
    "clave_prod_serv": ["clave_prod_serv", "clave", "claveprodsrv", "clave_producto", "clave_sat"],
    "descripcion": ["descripcion", "descripción", "desc", "concepto", "descripcion_concepto"],
    "cantidad": ["cantidad", "qty", "cant", "unidades"],
    "valor_unitario": [
        "valor_unitario",
        "precio",
        "precio_unitario",
        "valor",
        "total_concepto",  # se incluye por robustez
    ],
    "importe": ["importe", "total", "monto", "subtotal", "total_concepto"],
    "tasa": ["tasa", "tasa_iva", "porcentaje"],
}

# Construir diccionario inverso alias → canónico (en minúsculas)
for _canon, _alias_list in _ALIASES.items():
    for _alias in _alias_list:
        _ALIAS_MAP[_alias.lower()] = _canon


def _is_blank(value: Any) -> bool:
    """Detecta celdas vacías/NaN sin usar float como tipo de dato."""
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.lower() in ("nan", "none", "null")


def _to_decimal(value: Any) -> Decimal | None:
    """Convierte un valor a Decimal; retorna None si no es posible."""
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _to_decimal_or_error(
    canonical_vals: dict[str, Any],
    field: str,
    fila: int,
    archivo: str,
) -> Decimal | None:
    """Convierte una columna canónica a Decimal; lanza ValorNoNumericoError si no es posible."""
    raw = canonical_vals.get(field)
    if raw is None:
        return None
    result = _to_decimal(raw)
    if result is None:
        raise ValorNoNumericoError(
            f"Valor no numérico en columna '{field}', fila {fila}: {raw!r}",
            archivo=archivo,
        )
    return result


def _map_columns(column_names: list[str]) -> tuple[dict[str, str], list[str]]:
    """
    Retorna:
      - col_map: {columna_original → nombre_canónico}  solo para columnas mapeadas
      - unmapped: columnas originales sin mapear
    """
    col_map: dict[str, str] = {}
    unmapped: list[str] = []
    seen_canonicals: set[str] = set()

    for col in column_names:
        normalized = col.strip().lower()
        canonical = _ALIAS_MAP.get(normalized)
        if canonical and canonical not in seen_canonicals:
            col_map[col] = canonical
            seen_canonicals.add(canonical)
        else:
            unmapped.append(col)

    return col_map, unmapped


def _detect_csv_separator(ruta: Path) -> str:
    """Detecta si el CSV usa coma o punto y coma como separador."""
    try:
        first_line = ruta.read_text(encoding="utf-8").splitlines()[0]
    except UnicodeDecodeError:
        first_line = ruta.read_text(encoding="latin-1").splitlines()[0]

    semicolons = first_line.count(";")
    commas = first_line.count(",")
    return ";" if semicolons > commas else ","


def _load_csv(ruta: Path) -> tuple[list[dict[str, str]], list[AdvertenciaTabla]]:
    """Carga CSV como lista de dicts. Retorna (filas, advertencias)."""
    sep = _detect_csv_separator(ruta)
    try:
        text = ruta.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = ruta.read_text(encoding="latin-1")

    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    filas: list[dict[str, str]] = []
    for row in reader:
        if any(v and v.strip() for v in row.values()):
            filas.append({k: (v.strip() if v else "") for k, v in row.items()})
    return filas, []


def _load_xlsx(ruta: Path) -> tuple[list[dict[str, str]], str, list[AdvertenciaTabla]]:
    """Carga XLSX como lista de dicts. Retorna (filas, formato, advertencias)."""
    advertencias: list[AdvertenciaTabla] = []
    try:
        wb = load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as exc:
        raise ArchivoCorrumpidoError(
            f"XLSX inválido o corrupto: {exc}", archivo=str(ruta)
        ) from exc
    sheet_names = wb.sheetnames
    if len(sheet_names) > 1:
        advertencias.append(
            f"El archivo contiene {len(sheet_names)} hojas "
            f"({', '.join(str(s) for s in sheet_names)}); "
            f"se procesó únicamente la primera hoja: '{sheet_names[0]}'."
        )
    ws = wb[sheet_names[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return [], "xlsx", advertencias
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    filas: list[dict[str, str]] = []
    for row in rows_iter:
        vals = [str(v).strip() if v is not None else "" for v in row]
        if any(v for v in vals):
            filas.append(dict(zip(headers, vals, strict=False)))
    wb.close()
    return filas, "xlsx", advertencias


def _load_xls(ruta: Path) -> tuple[list[dict[str, str]], str, list[AdvertenciaTabla]]:
    """Carga XLS (formato antiguo) como lista de dicts. Retorna (filas, formato, advertencias)."""
    advertencias: list[AdvertenciaTabla] = []
    try:
        wb = xlrd.open_workbook(str(ruta))
    except Exception as exc:
        raise ArchivoCorrumpidoError(
            f"XLS inválido o corrupto: {exc}", archivo=str(ruta)
        ) from exc
    sheet_names = wb.sheet_names()
    if len(sheet_names) > 1:
        advertencias.append(
            f"El archivo contiene {len(sheet_names)} hojas "
            f"({', '.join(str(s) for s in sheet_names)}); "
            f"se procesó únicamente la primera hoja: '{sheet_names[0]}'."
        )
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return [], "xls", advertencias
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    filas: list[dict[str, str]] = []
    for r in range(1, ws.nrows):
        vals: list[str] = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                val = str(int(cell.value)) if cell.value == int(cell.value) else str(cell.value)
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                val = ""
            else:
                val = str(cell.value).strip()
            vals.append(val)
        if any(v for v in vals):
            filas.append(dict(zip(headers, vals, strict=False)))
    return filas, "xls", advertencias


def _load_ods(ruta: Path) -> tuple[list[dict[str, str]], str, list[AdvertenciaTabla]]:
    """Carga ODS como lista de dicts usando lxml+zipfile. Retorna (filas, formato, advertencias)."""
    advertencias: list[AdvertenciaTabla] = []
    try:
        with zipfile.ZipFile(str(ruta)) as zf:
            content_xml = zf.read("content.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ArchivoCorrumpidoError(f"ODS inválido o corrupto: {exc}", archivo=str(ruta)) from exc

    root = etree_fromstring(content_xml)
    tables = root.findall(f".//{{{_NS_TABLE}}}table")

    if not tables:
        return [], "ods", advertencias

    if len(tables) > 1:
        sheet_names = [t.get(f"{{{_NS_TABLE}}}name", f"Hoja{i + 1}") for i, t in enumerate(tables)]
        advertencias.append(
            f"El archivo contiene {len(tables)} hojas "
            f"({', '.join(sheet_names)}); "
            f"se procesó únicamente la primera hoja: '{sheet_names[0]}'."
        )

    table = tables[0]
    raw_rows: list[list[str]] = []

    for row_el in table.findall(f"{{{_NS_TABLE}}}table-row"):
        cells: list[str] = []
        for cell_el in row_el.findall(f"{{{_NS_TABLE}}}table-cell"):
            repeat = int(cell_el.get(f"{{{_NS_TABLE}}}number-columns-repeated", "1"))
            texts = cell_el.findall(f".//{{{_NS_TEXT}}}p")
            val = "".join(t.text or "" for t in texts).strip()
            cells.extend([val] * repeat)
        # Recortar celdas vacías repetidas al final (artefacto del formato)
        while cells and not cells[-1]:
            cells.pop()
        raw_rows.append(cells)

    # Eliminar filas vacías al final
    while raw_rows and not any(raw_rows[-1]):
        raw_rows.pop()

    if not raw_rows:
        return [], "ods", advertencias

    headers = raw_rows[0]
    filas: list[dict[str, str]] = []
    for row_vals in raw_rows[1:]:
        if any(v for v in row_vals):
            row_dict = {
                headers[i]: (row_vals[i] if i < len(row_vals) else "") for i in range(len(headers))
            }
            filas.append(row_dict)
    return filas, "ods", advertencias


def _load_dataframe(ruta: Path) -> tuple[list[dict[str, str]], str, list[AdvertenciaTabla]]:
    """Carga filas como lista de dicts {col_name: value_str}.

    Retorna (filas, formato, advertencias).
    - filas: cada elemento es un dict {header: valor_string}
    - formato: "csv", "xlsx", "xls" u "ods"
    - advertencias: lista de strings (multi-hoja, encoding, etc.)

    Filas completamente vacías NO se incluyen (ya filtradas aquí).
    """
    ext = ruta.suffix.lower()

    if ext == ".csv":
        filas, advertencias = _load_csv(ruta)
        return filas, "csv", advertencias

    elif ext == ".xlsx":
        return _load_xlsx(ruta)

    elif ext == ".xls":
        return _load_xls(ruta)

    elif ext == ".ods":
        return _load_ods(ruta)

    else:
        raise FormatoNoSoportadoError(
            f"Extensión '{ext}' no soportada. Use .csv, .xlsx, .xls u .ods.",
            archivo=str(ruta),
        )


def parsear_tabla(ruta: Path) -> TablaResult:
    """
    Parsea un archivo CSV o XLSX con una tabla de conceptos.

    Returns:
        TablaResult con los registros normalizados.

    Raises:
        ArchivoNoEncontradoError: si el archivo no existe (exit 3).
        FormatoNoSoportadoError: si la extensión no es .csv/.xlsx (exit 1).
        ArchivoSinDatosError: si el archivo está vacío o sin filas de datos (exit 1).
        ColumnaRequeridaAusenteError: si falta una columna requerida (exit 1).
        ValorNoNumericoError: si un campo monetario tiene valor no numérico (exit 1).
    """
    if not ruta.exists():
        raise ArchivoNoEncontradoError(
            f"Archivo no encontrado: {ruta}",
            archivo=str(ruta),
        )

    filas, formato, advertencias_carga = _load_dataframe(ruta)

    # Detectar archivo sin datos
    if not filas:
        raise ArchivoSinDatosError(
            f"El archivo no contiene filas de datos: {ruta}",
            archivo=str(ruta),
        )

    # Mapear columnas usando los nombres del primer dict
    col_names = list(filas[0].keys())
    col_map, unmapped = _map_columns(col_names)
    columnas_detectadas = list(col_map.values())

    # Verificar columnas requeridas
    columnas_encontradas = set(col_map.values())
    for col_requerida in sorted(_COLUMNAS_REQUERIDAS):
        if col_requerida not in columnas_encontradas:
            raise ColumnaRequeridaAusenteError(
                f"Columna requerida no encontrada: '{col_requerida}'",
                archivo=str(ruta),
            )

    registros: list[TablaRow] = []

    tasa_presente = "tasa" in columnas_detectadas
    total_iva: Decimal | None = Decimal("0") if tasa_presente else None

    for fila_idx, row_dict in enumerate(filas, start=2):  # fila 1 es el header
        # Columnas canónicas
        canonical_vals: dict[str, Any] = {}
        for orig_col, canon in col_map.items():
            val: Any = row_dict.get(orig_col, "") or None  # "" → None
            if _is_blank(val):
                val = None
            canonical_vals[canon] = val

        # Extras
        columnas_extra: dict[str, Any] = {}
        for col in unmapped:
            val = row_dict.get(col, "") or None
            if _is_blank(val):
                val = None
            columnas_extra[col] = val

        # Convertir decimales — valor no numérico es error fatal
        archivo_str = str(ruta)
        cantidad = _to_decimal_or_error(canonical_vals, "cantidad", fila_idx, archivo_str)
        valor_unitario = _to_decimal_or_error(
            canonical_vals, "valor_unitario", fila_idx, archivo_str
        )
        importe = _to_decimal_or_error(canonical_vals, "importe", fila_idx, archivo_str)

        # Feature 3: IVA estimado por fila
        tasa_raw = canonical_vals.get("tasa") if tasa_presente else None
        iva_estimado: Decimal | None = None
        if tasa_presente and tasa_raw is not None:
            tasa_decimal = _to_decimal(tasa_raw)
            if tasa_decimal is None:
                advertencias_carga.append(
                    AdvertenciaTasaNoNumerica(
                        fila=fila_idx,
                        valor_encontrado=str(tasa_raw),
                    )
                )
            elif importe is not None:
                iva_estimado = importe * tasa_decimal

        registro = TablaRow(
            fila=fila_idx,
            clave_prod_serv=canonical_vals.get("clave_prod_serv"),
            descripcion=canonical_vals.get("descripcion"),
            cantidad=cantidad,
            valor_unitario=valor_unitario,
            importe=importe,
            tasa=tasa_raw if tasa_presente else None,
            iva_estimado=iva_estimado,
            columnas_extra=columnas_extra,
        )
        if tasa_presente:
            registro._tasa_col_presente = True
        registros.append(registro)

        if iva_estimado is not None and total_iva is not None:
            total_iva += iva_estimado

    # Feature 2: validación cruzada de importes
    for registro in registros:
        if (
            registro.cantidad is not None
            and registro.valor_unitario is not None
            and registro.importe is not None
        ):
            esperado = registro.cantidad * registro.valor_unitario
            diferencia = abs(registro.importe - esperado)
            if diferencia > Decimal("0.01"):
                advertencias_carga.append(
                    AdvertenciaImporteInconsistente(
                        fila=registro.fila,
                        importe_declarado=str(registro.importe),
                        importe_calculado=f"{esperado:.2f}",
                        diferencia=f"{diferencia:.2f}",
                    )
                )

    return TablaResult(
        archivo=str(ruta.resolve()),
        formato=formato,
        total_registros=len(registros),
        columnas_detectadas=columnas_detectadas,
        registros=registros,
        total_iva_estimado=total_iva,
        advertencias=advertencias_carga,
    )
