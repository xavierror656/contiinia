"""Generador de reporte Excel para `contiinia exportar`."""

from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from contiinia.errors import (
    ContiiniaError,
    DirectorioNoEncontradoError,
    ErrorEscrituraError,
    FormatoNoSoportadoError,
    PermisoDenegadoError,
    RutaNoEsDirectorioError,
)
from contiinia.models.exportar import ExportarPeriodo, ExportarResult
from contiinia.parsers.xml import parsear_xml


def generar_exportar(directorio: Path, salida: Path, recursivo: bool = False) -> ExportarResult:
    """Procesa CFDIs en *directorio* y genera un reporte Excel en *salida*.

    Raises:
        DirectorioNoEncontradoError: si el directorio no existe (exit 3).
        RutaNoEsDirectorioError: si la ruta no es un directorio (exit 3).
        FormatoNoSoportadoError: si salida no termina en .xlsx (exit 1).
        ErrorEscrituraError: si no se puede escribir el archivo de salida (exit 3).
    """
    if not directorio.exists():
        raise DirectorioNoEncontradoError(
            f"Directorio no encontrado: {directorio}",
            archivo=str(directorio),
        )
    if not directorio.is_dir():
        raise RutaNoEsDirectorioError(
            f"La ruta no es un directorio: {directorio}",
            archivo=str(directorio),
        )
    if salida.suffix.lower() != ".xlsx":
        raise FormatoNoSoportadoError(
            f"El archivo de salida debe tener extensión .xlsx, se recibió: '{salida.suffix}'",
            archivo=str(salida),
        )

    padre = salida.parent
    if not padre.exists() or not padre.is_dir():
        raise ErrorEscrituraError(
            f"El directorio padre de la salida no existe o no es accesible: {padre}",
            archivo=str(salida),
        )

    pattern = "**/*.xml" if recursivo else "*.xml"
    try:
        archivos = sorted(directorio.glob(pattern))
    except PermissionError as exc:
        raise PermisoDenegadoError(
            f"Sin permiso para leer el directorio: {exc}",
            archivo=str(directorio),
        ) from exc

    cfdi_exitosos = []
    errores_lista: list[dict[str, str]] = []

    for archivo in archivos:
        ruta_str = str(archivo)
        try:
            cfdi = parsear_xml(archivo)
            cfdi_exitosos.append(cfdi)
        except ContiiniaError as exc:
            errores_lista.append({
                "archivo": ruta_str,
                "error": exc.error_type,
                "detalle": exc.detalle,
            })
        except Exception as exc:
            errores_lista.append({
                "archivo": ruta_str,
                "error": "error_inesperado",
                "detalle": str(exc),
            })

    # --- Calcular totales del resumen a partir de los CFDIs ya parseados ---
    subtotal_i = Decimal("0")
    subtotal_e = Decimal("0")
    total_i = Decimal("0")
    total_e = Decimal("0")
    iva_importe: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    total_iva_retenido = Decimal("0")
    total_isr_retenido = Decimal("0")
    conteo_tipo: dict[str, int] = defaultdict(int)
    monedas_set: set[str] = set()
    fechas: list[str] = []

    for cfdi in cfdi_exitosos:
        tipo = cfdi.tipo_de_comprobante
        conteo_tipo[tipo] += 1

        moneda = cfdi.moneda or "MXN"
        monedas_set.add(moneda)

        if cfdi.fecha:
            fechas.append(cfdi.fecha)

        # Factor de conversión MXN (igual que resumen.py)
        tc_factor = Decimal("1")
        moneda_excluida = False
        if moneda != "MXN" and tipo in ("I", "E"):
            tc = cfdi.tipo_cambio
            if tc is None or tc == Decimal("0"):
                moneda_excluida = True
            else:
                tc_factor = tc

        traslados_g = cfdi.impuestos.traslados if cfdi.impuestos else []
        retenciones_g = cfdi.impuestos.retenciones if cfdi.impuestos else []

        subtotal_orig = cfdi.subtotal or Decimal("0")
        total_orig = cfdi.total or Decimal("0")

        if tipo == "I" and not moneda_excluida:
            subtotal_i += subtotal_orig * tc_factor
            total_i += total_orig * tc_factor
            for t in traslados_g:
                if t.impuesto == "002":
                    clave = (t.tipo_factor or "Tasa", str(t.tasa_o_cuota or ""))
                    iva_importe[clave] += (t.importe or Decimal("0")) * tc_factor
            for r in retenciones_g:
                if r.impuesto == "002":
                    total_iva_retenido += (r.importe or Decimal("0")) * tc_factor
                elif r.impuesto == "001":
                    total_isr_retenido += (r.importe or Decimal("0")) * tc_factor
        elif tipo == "E" and not moneda_excluida:
            subtotal_e += subtotal_orig * tc_factor
            total_e += total_orig * tc_factor
            for t in traslados_g:
                if t.impuesto == "002":
                    clave = (t.tipo_factor or "Tasa", str(t.tasa_o_cuota or ""))
                    iva_importe[clave] -= (t.importe or Decimal("0")) * tc_factor

    total_neto = total_i - total_e
    total_iva_trasladado = sum(
        v for k, v in iva_importe.items() if k[0] != "Exento"
    ) or Decimal("0")
    monedas_lista = sorted(monedas_set, key=lambda m: (0 if m == "MXN" else 1, m))
    fecha_min = min(fechas) if fechas else None
    fecha_max = max(fechas) if fechas else None

    # --- Construir Excel ---
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    # Hoja 1: Resumen
    ws_res = wb.create_sheet("Resumen")
    for clave, valor in [
        ("subtotal_ingresos", str(subtotal_i)),
        ("subtotal_egresos", str(subtotal_e)),
        ("total_neto", str(total_neto)),
        ("total_iva_trasladado", str(total_iva_trasladado)),
        ("total_iva_retenido", str(total_iva_retenido)),
        ("total_isr_retenido", str(total_isr_retenido)),
        ("fecha_min", fecha_min or ""),
        ("fecha_max", fecha_max or ""),
        ("count_I", conteo_tipo.get("I", 0)),
        ("count_E", conteo_tipo.get("E", 0)),
        ("count_P", conteo_tipo.get("P", 0)),
        ("count_N", conteo_tipo.get("N", 0)),
        ("count_T", conteo_tipo.get("T", 0)),
        ("monedas_detectadas", ", ".join(monedas_lista)),
    ]:
        ws_res.append([clave, valor])

    # Hoja 2: Detalle_CFDI (valores en moneda original del comprobante — QA-EXP-01)
    ws_det = wb.create_sheet("Detalle_CFDI")
    ws_det.append([
        "uuid", "tipo", "fecha", "emisor_rfc", "emisor_nombre",
        "receptor_rfc", "receptor_nombre", "moneda",
        "subtotal", "total", "metodo_pago", "forma_pago",
    ])
    for cfdi in cfdi_exitosos:
        ws_det.append([
            cfdi.uuid,
            cfdi.tipo_de_comprobante,
            cfdi.fecha,
            cfdi.emisor.rfc,
            cfdi.emisor.nombre or "",
            cfdi.receptor.rfc,
            cfdi.receptor.nombre or "",
            cfdi.moneda or "MXN",
            str(cfdi.subtotal or Decimal("0")),
            str(cfdi.total or Decimal("0")),
            cfdi.metodo_pago or "",
            cfdi.forma_pago or "",
        ])

    # Hoja 3: Errores
    ws_err = wb.create_sheet("Errores")
    ws_err.append(["archivo", "error", "detalle"])
    for e in errores_lista:
        ws_err.append([e["archivo"], e["error"], e["detalle"]])

    try:
        wb.save(str(salida))
    except (PermissionError, OSError) as exc:
        raise ErrorEscrituraError(
            f"No se pudo escribir el archivo de salida: {exc}",
            archivo=str(salida),
        ) from exc

    return ExportarResult(
        archivo_generado=str(salida.resolve()),
        hojas=["Resumen", "Detalle_CFDI", "Errores"],
        total_cfdi_exitosos=len(cfdi_exitosos),
        total_cfdi_errores=len(errores_lista),
        periodo=ExportarPeriodo(fecha_min=fecha_min, fecha_max=fecha_max),
    )
